import os
import django

# Set up Django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "interview_task_by_django.settings")
django.setup()

# Import your models here
from main_app.models import Employee
from openpyxl import load_workbook
from datetime import datetime


def fill_db_by_excel_data():
    workbook = load_workbook('Employee-table.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        date = row[3]
        if not isinstance(date, datetime):
            date = datetime.strptime(date, '%d/%m/%Y')

        Employee.objects.create(
            first_name=row[0],
            last_name=row[1],
            mobile=int(row[2]),
            start_date=date,
            position=row[4],
            salary=row[5],
            employee_id=row[6]
        )


fill_db_by_excel_data()
